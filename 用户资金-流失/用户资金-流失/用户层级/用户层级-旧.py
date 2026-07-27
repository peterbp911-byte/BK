import os
import time
import random
from datetime import datetime, timedelta

import openpyxl
import requests
import urllib3
from bs4 import BeautifulSoup

import platfrom_config
from platfrom_config import get_platfrom_cookies

# 禁用 InsecureRequestWarning
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

yesterday = (datetime.today() - timedelta(days=1)).strftime("%Y-%m-%d")
yesterday_folder = (datetime.today() - timedelta(days=1)).strftime("%m%d")

CSV_DIR = r"D:\TT用户分层"
platform_list = ['7ss','vana7','7luck','brlucky','7xx','7aa','novo7', 'b7', 'sp7', 'b777', '1xspin', 'spin77', 'sp1', 'bx365', 'brplay7', 'gana7', 'brslot', '7pg', 'brspin', 'brwins', 'x7s']

for platform_name in platform_list:
    BASE_URL = platfrom_config.get_platfrom_config(platform_name)["LE_URL"]

    search_params = {
        '注册当日未充值': 'reg_curday_unpay',
        '首充次日未登录': 'first_nextday_unlogin',
        '首充3日未登录': 'first_3d_unlogin',
        '首充7日未登录': 'first_7d_unlogin',
        '首充当日亏损用户': 'first_curday_loss'
    }

    # === 会话 ===
    session = requests.Session()
    session.headers.update({
        "accept": "text/html, application/xhtml+xml",
        "accept-language": "zh-CN,zh;q=0.9",
        "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/143.0.0.0 Safari/537.36",
        "referer": BASE_URL,
    })

    cookies = get_platfrom_cookies(platform_name)
    if not cookies:
        print(f"❌ 未获取到Cookie: {platform_name}")
        exit(1)

    session.cookies.update(cookies)

    FIELD_MAP = {
        "UID": "UID",
        "平台名称": "平台名称",
        "用户名": "用户名",
        "手机号": "手机号",
        "注册时间": "注册时间",
        "最后登录时间": "最后登录时间"
    }

    # ============================================
    def fetch_page(page, search_key, search_value, max_retries=5):
        params = {
            "search": {search_value: search_value},
            "day": yesterday,
            "size": 500,
            "page": page,
        }

        for attempt in range(1, max_retries + 1):
            try:
                res = session.get(
                    BASE_URL,
                    params=params,
                    timeout=30,
                    verify=False
                )

                if res.status_code != 200:
                    print(f"页面{page}状态码{res.status_code}，重试{attempt}")
                    time.sleep(2 ** attempt + random.uniform(0, 1))
                    continue

                soup = BeautifulSoup(res.text, "lxml")

                thead = soup.find("thead")
                if not thead:
                    print(f"页面{page}找不到表头，重试{attempt}")
                    time.sleep(1)
                    continue

                raw_titles = [th.get_text(strip=True) for th in thead.find_all("th")]

                titles_page = []
                for t in raw_titles:
                    titles_page.append(FIELD_MAP.get(t, t.replace(" ", "_").lower()))

                tbody = soup.find("tbody")
                if not tbody:
                    print(f"页面{page}找不到表体，重试{attempt}")
                    time.sleep(1)
                    continue

                rows = []
                for tr in tbody.find_all("tr"):
                    cols = [td.get_text(strip=True) for td in tr.find_all("td")]
                    if cols:
                        rows.append(cols)

                if not rows:
                    print(f"页面{page}无数据行，重试{attempt}")
                    time.sleep(1)
                    continue

                # 页数解析
                page_info = soup.find("span", string=lambda x: x and "第" in x and "页" in x)
                if page_info:
                    text = page_info.get_text(strip=True)
                    part = text.split(" ")[0]
                    now_page, total_pages = part.replace("第", "").replace("页", "").split("/")
                    total_records = int(text.split("共")[1].replace("条记录", "").strip())
                else:
                    total_pages = 1
                    total_records = len(rows)

                return titles_page, rows, int(total_pages), total_records

            except (ConnectionResetError, requests.exceptions.ConnectionError) as e:
                print(f"页面{page}连接错误: {e}，重试{attempt}/{max_retries}")
                time.sleep(3 ** attempt + random.uniform(0, 2))
            except Exception as e:
                print(f"页面{page}异常: {e}，重试{attempt}")
                time.sleep(2 ** attempt + random.uniform(0, 1))

        print(f"页面{page}最终失败，跳过")
        return None, [], 1, 0

    # ============================================
    def fetch_and_parse_for_search(search_key, search_value):
        print(f"\n=== 开始抓取数据: {platform_name}  {search_key} ===")
        all_rows = []
        failed_pages = []

        titles_page, rows, total_pages, total_records = fetch_page(1, search_key, search_value)

        # 如果连表头都拿不到，直接返回空列表
        if titles_page is None:
            print(f"→ 无法获取表头，跳过 {search_key}")
            return []

        print(f"记录 {total_records} 条，共 {total_pages} 页")

        if rows:
            all_rows.extend(rows)
        else:
            failed_pages.append(1)

        # 批量抓取所有页面
        for page in range(2, total_pages + 1):
            time.sleep(0.3)  # 降低请求频率
            _, rows, _, _ = fetch_page(page, search_key, search_value)
            if rows:
                all_rows.extend(rows)
            else:
                failed_pages.append(page)
            if page % 100 == 0:
                print(f"已处理{page}页，当前数据{len(all_rows)}条")

        # 重试失败页面
        if failed_pages:
            print(f"\n重试{len(failed_pages)}个失败页面...")
            for page in failed_pages:
                time.sleep(1)
                _, rows, _, _ = fetch_page(page, search_key, search_value)
                if rows:
                    all_rows.extend(rows)
                    print(f"页面{page}重试成功")

        if not all_rows:
            print("→ 无记录")
            return []

        # 提取 UID、手机号、平台名称
        uid_idx = phone_idx = plat_idx = None
        for i, col in enumerate(titles_page):
            if col == "UID":
                uid_idx = i
            elif col == "手机号":
                phone_idx = i
            elif col == "平台名称":
                plat_idx = i

        if uid_idx is None or phone_idx is None or plat_idx is None:
            print("→ 未找到 UID、手机号 或 平台名称 列")
            return []

        filtered_rows = []
        for row in all_rows:
            filtered_rows.append([row[uid_idx], row[phone_idx], row[plat_idx]])

        return filtered_rows

    # ============================================
    def main():
        excel_dir = os.path.join(CSV_DIR, yesterday_folder, platform_name)
        os.makedirs(excel_dir, exist_ok=True)
        excel_file = os.path.join(excel_dir, f"{platform_name}_分层数据_{yesterday_folder}.xlsx")
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        for search_key, search_value in search_params.items():
            filtered_rows = fetch_and_parse_for_search(search_key, search_value)
            if filtered_rows:
                ws = wb.create_sheet(title=search_key)
                ws.append(["UID", "手机号", "平台名称"])
                for row in filtered_rows:
                    ws.append(row)
                for col in ws.columns:
                    for cell in col:
                        cell.number_format = '@'

        if wb.sheetnames:
            wb.save(excel_file)
            print(f"\n→ 所有分层已写入 {excel_file}")
        else:
            print(f"\n→ {platform_name} 所有分层均无数据，跳过保存")
        print("\n=== 完成 ===")

    if __name__ == "__main__":
        main()