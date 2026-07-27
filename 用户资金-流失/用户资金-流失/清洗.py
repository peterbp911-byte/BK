import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Alignment, PatternFill, Border, Side
import datetime
import os

yesterday = (datetime.datetime.now() - datetime.timedelta(days=1)).strftime('%m%d')
platfrom_name = ['b777','brl77','1xspin','hot77','spin77','sp1','viva7','bx365','brplay7','gana7','brslot','7pg','brspin','brwins','x7s']

for platfrom in platfrom_name:
    def classify_recharge(amount):
        if amount > 5000:
            return '>5000'
        elif amount > 3000:
            return '>3000 <=5000'
        elif amount > 1000:
            return '>1000 <=3000'
        elif amount > 500:
            return '>500 <=1000'
        elif amount > 200:
            return '>200 <=500'
        elif amount >= 100:
            return '>=100 <=200'
        else:
            return '其他'

def sanitize_filename(filename):
    """Replace invalid filename characters with safe alternatives"""
    invalid_chars = {'<': '(', '>': ')', '|': '-', ':': '-', '"': "'", '?': ''}
    for char, replacement in invalid_chars.items():
        filename = filename.replace(char, replacement)
    return filename

for platfrom in platfrom_name:
    file_path = os.path.join('D:\TT用户资金数据', yesterday, platfrom, f'{platfrom}_user_funds.csv')
    if not os.path.exists(file_path):
        print(f'{platfrom} 数据文件不存在，跳过处理。\n')  
        continue

    try:
        df = pd.read_csv(file_path, encoding='utf-8')
    except UnicodeDecodeError:
        df = pd.read_csv(file_path, encoding='gbk')

    #统一格式
    df.rename(columns={'历史充值_⥮': '历史充值'}, inplace=True)
    df.rename(columns= {'历史充值_?': '历史充值'}, inplace=True)
    df.rename(columns={'UID': '用户uid'}, inplace=True)
    df['用户uid'] = df['用户uid'].astype(str)
    df['手机号'] = df['手机号'].astype(str)
    df['历史充值'] = pd.to_numeric(df['历史充值'], errors='coerce').fillna(0)
    
    df['充值等级'] = df['历史充值'].apply(classify_recharge)
    
    # 按充值等级分组输出
    categories = ['>5000', '>3000 <=5000', '>1000 <=3000', '>500 <=1000', '>200 <=500', '>=100 <=200']
    
    output_dir = os.path.join('D:\TT用户资金数据', yesterday, platfrom)
    os.makedirs(output_dir, exist_ok=True)
    
    # 创建Excel文件
    output_file = os.path.join(output_dir, f'充值分类_{platfrom}.xlsx')
    wb = openpyxl.Workbook()
    ws = wb.active
    
    col = 1  # 起始列
    for category in categories:
        # 过滤对应等级的数据
        category_df = df[df['充值等级'] == category][['用户uid', '手机号']]
        
        if len(category_df) > 0:
            # 写入分类标题
            ws.cell(row=1, column=col, value=category)
            ws.cell(row=1, column=col+1, value='')
            
            # 写入数据
            for idx, (_, row) in enumerate(category_df.iterrows(), start=2):
                cell_uid = ws.cell(row=idx, column=col, value=row['用户uid'])
                cell_phone = ws.cell(row=idx, column=col+1, value=row['手机号'])
                # 设置为文本格式
                cell_uid.number_format = '@'
                cell_phone.number_format = '@'
            
            print(f'{platfrom} - {category}: {len(category_df)}条数据')
            col += 5  # 间隔3列（当前2列+3列空白=5列）
    
    wb.save(output_file)
    print(f'{platfrom} 分类完成！数据已输出到 {output_file}\n')





