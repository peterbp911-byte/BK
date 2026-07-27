import os
from datetime import datetime, timedelta

import pandas as pd
import requests
from sqlalchemy import create_engine
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


TOKEN = "8718125099:AAH0lKs45qIyZUuAmWL8iE2NMh7H-lZ5hhI"
SEND_ID = ["-5255411609"]

SEND_AS_IMAGE = True
SEND_IMAGE_AS_DOCUMENT = True
IMAGE_DPI = 360

today = datetime.today()
month_start_str = today.replace(day=1).strftime("%Y-%m-%d")
yesterday_str = (today - timedelta(days=1)).strftime("%Y-%m-%d")
date_str = today.strftime("%Y%m%d")

output_dir = "C:/日报存档"
os.makedirs(output_dir, exist_ok=True)

engine = create_engine("postgresql+psycopg2://postgres:147258@localhost:5432/postgres")

# 15日留存在业务口径里展示为14日留存。
RETENTION_WINDOWS = [
	(2, "2日"),
	(3, "3日"),
	(4, "4日"),
	(7, "7日"),
	(15, "14日"),
	(30, "30日"),
]


def query_recharge_retention(plat: str) -> pd.DataFrame:
	select_fields = [
		'"statisDate"::date AS 日期',
		'COALESCE("firstDepositUserNum", 0) AS 首充人数',
		'COALESCE("rechageAmt", 0) AS 首充金额',
	]

	for day_num, label in RETENTION_WINDOWS:
		select_fields.extend([
			f'COALESCE("rechageAmt{day_num}", 0) AS "{label}充值金额"',
			f'COALESCE("retainRatio{day_num}", 0) / 100.0 AS "{label}充值留存率"',
			f'COALESCE("retainUserNum{day_num}", 0) AS "{label}留存人数"',
		])

	sql = f"""
	SELECT
		{', '.join(select_fields)}
	FROM "public"."{plat}_ReRetaintion"
	WHERE "statisDate" BETWEEN '{month_start_str}'::date AND '{yesterday_str}'::date
	ORDER BY "statisDate" DESC
	"""
	return pd.read_sql(sql, engine)


def query_bet_retention(plat: str) -> pd.DataFrame:
	select_fields = [
		'"statisDate"::date AS 日期',
		'COALESCE("newBetUserNum", 0) AS 首投人数',
		'COALESCE("betAmt", 0) AS 首投金额',
	]

	for day_num, label in RETENTION_WINDOWS:
		select_fields.extend([
			f'COALESCE("betAmt{day_num}", 0) AS "{label}投注金额"',
			f'COALESCE("retainRatio{day_num}", 0) / 100.0 AS "{label}投注留存率"',
			f'COALESCE("retainUserNum{day_num}", 0) AS "{label}留存人数"',
		])

	sql = f"""
	SELECT
		{', '.join(select_fields)}
	FROM "public"."{plat}_BetRetaintion"
	WHERE "statisDate" BETWEEN '{month_start_str}'::date AND '{yesterday_str}'::date
	ORDER BY "statisDate" DESC
	"""
	return pd.read_sql(sql, engine)


def _to_numeric(series: pd.Series) -> pd.Series:
	cleaned = series.astype(str).str.replace("%", "", regex=False).str.replace(",", "", regex=False)
	return pd.to_numeric(cleaned, errors="coerce")


def add_summary_rows(df: pd.DataFrame) -> pd.DataFrame:
	if df.empty:
		return df

	numeric_cols = [c for c in df.columns if c != "日期"]
	numeric_data = {c: _to_numeric(df[c]) for c in numeric_cols}
	day_count = len(df)

	summary_row = {c: "" for c in df.columns}
	average_row = {c: "" for c in df.columns}
	summary_row["日期"] = "汇总"
	average_row["日期"] = "平均"

	for col in numeric_cols:
		col_sum = float(numeric_data[col].sum(skipna=True))
		col_avg = float(numeric_data[col].mean(skipna=True)) if day_count else 0

		if col.endswith("留存率"):
			summary_row[col] = round(col_avg, 4)
			average_row[col] = round(col_avg, 4)
		else:
			summary_row[col] = round(col_sum, 4)
			average_row[col] = round(col_sum / day_count, 4) if day_count else 0

	return pd.concat([pd.DataFrame([summary_row, average_row]), df], ignore_index=True)


def style_retention_report(excel_path: str):
	wb = load_workbook(excel_path)

	header_fill = PatternFill("solid", fgColor="0F766E")
	header_font = Font(color="FFFFFF", bold=True)
	summary_fill = PatternFill("solid", fgColor="DDF3EE")
	thin = Side(style="thin", color="A9CFC6")
	border = Border(left=thin, right=thin, top=thin, bottom=thin)

	for ws in wb.worksheets:
		for col_idx in range(1, ws.max_column + 1):
			cell = ws.cell(row=1, column=col_idx)
			cell.fill = header_fill
			cell.font = header_font
			cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
			cell.border = border

		for row_idx in range(2, ws.max_row + 1):
			row_label = ws.cell(row=row_idx, column=1).value
			is_summary = str(row_label) in {"汇总", "平均"}

			for col_idx in range(1, ws.max_column + 1):
				cell = ws.cell(row=row_idx, column=col_idx)
				col_name = ws.cell(row=1, column=col_idx).value
				cell.border = border
				cell.alignment = Alignment(horizontal="center", vertical="center")

				if is_summary:
					cell.fill = summary_fill
					cell.font = Font(bold=True)

				if isinstance(cell.value, (int, float)) and pd.notna(cell.value) and col_name:
					if str(col_name).endswith("留存率"):
						cell.number_format = "0.0%"
					else:
						cell.number_format = "#,##0"

		for col_idx in range(1, ws.max_column + 1):
			max_len = 0
			for row_idx in range(1, ws.max_row + 1):
				value = ws.cell(row=row_idx, column=col_idx).value
				if value is None:
					continue
				max_len = max(max_len, len(str(value)))
			ws.column_dimensions[get_column_letter(col_idx)].width = min(max(12, max_len + 4), 36)

		ws.row_dimensions[1].height = 34
		for row_idx in range(2, ws.max_row + 1):
			ws.row_dimensions[row_idx].height = 20

		ws.freeze_panes = "A2"
		ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
		ws.page_setup.paperSize = ws.PAPERSIZE_A4
		ws.sheet_properties.pageSetUpPr.fitToPage = True
		ws.page_setup.fitToWidth = 1
		ws.page_setup.fitToHeight = 0
		ws.print_area = ws.dimensions

	wb.save(excel_path)


def export_excel_to_pdf(excel_path: str):
	pdf_path = os.path.splitext(excel_path)[0] + ".pdf"
	excel_app = None
	workbook = None
	try:
		import win32com.client

		excel_app = win32com.client.DispatchEx("Excel.Application")
		excel_app.Visible = False
		excel_app.DisplayAlerts = False
		workbook = excel_app.Workbooks.Open(os.path.abspath(excel_path))

		for sheet in workbook.Worksheets:
			sheet.PageSetup.Zoom = False
			sheet.PageSetup.Orientation = 2
			sheet.PageSetup.FitToPagesWide = 1
			sheet.PageSetup.FitToPagesTall = False

		workbook.ExportAsFixedFormat(0, os.path.abspath(pdf_path))
		return pdf_path
	except Exception as e:
		print(f"PDF导出失败: {e}")
		return None
	finally:
		if workbook is not None:
			workbook.Close(False)
		if excel_app is not None:
			excel_app.Quit()


def convert_pdf_pages_to_images(pdf_path: str, dpi: int = 360, max_pages: int = 2):
	if not os.path.exists(pdf_path):
		return []

	image_paths = []
	try:
		import fitz

		zoom = dpi / 72
		doc = fitz.open(pdf_path)
		try:
			page_count = min(doc.page_count, max_pages)
			for idx in range(page_count):
				page = doc.load_page(idx)
				blocks = page.get_text("blocks")

				if blocks:
					x0 = min(b[0] for b in blocks)
					y0 = min(b[1] for b in blocks)
					x1 = max(b[2] for b in blocks)
					y1 = max(b[3] for b in blocks)
					margin = 8
					clip = fitz.Rect(
						max(page.rect.x0, x0 - margin),
						max(page.rect.y0, y0 - margin),
						min(page.rect.x1, x1 + margin),
						min(page.rect.y1, y1 + margin),
					)
				else:
					clip = page.rect

				pix = page.get_pixmap(matrix=fitz.Matrix(zoom, zoom), clip=clip, alpha=False)
				image_path = os.path.splitext(pdf_path)[0] + f"_p{idx + 1}.png"
				pix.save(image_path)
				image_paths.append(image_path)
		finally:
			doc.close()
	except Exception as e:
		print(f"PDF转图片失败: {e}")

	return image_paths


def send_telegram_document(chat_id: str, file_path: str, caption: str = "") -> bool:
	if not os.path.exists(file_path):
		print(f"待发送文件不存在: {file_path}")
		return False

	send_url = f"https://api.telegram.org/bot{TOKEN}/sendDocument"
	try:
		with open(file_path, "rb") as f:
			response = requests.post(
				send_url,
				data={"chat_id": chat_id, "caption": caption},
				files={"document": f},
				timeout=60,
			)
		if response.status_code == 200:
			return True
		print(f"Telegram发送失败({chat_id})，状态码: {response.status_code}，响应: {response.text}")
		return False
	except Exception as e:
		print(f"Telegram发送异常({chat_id}): {e}")
		return False


def send_telegram_photo(chat_id: str, image_path: str, caption: str = "") -> bool:
	if not os.path.exists(image_path):
		print(f"待发送图片不存在: {image_path}")
		return False

	send_url = f"https://api.telegram.org/bot{TOKEN}/sendPhoto"
	try:
		with open(image_path, "rb") as f:
			response = requests.post(
				send_url,
				data={"chat_id": chat_id, "caption": caption},
				files={"photo": f},
				timeout=60,
			)
		if response.status_code == 200:
			return True
		print(f"Telegram图片发送失败({chat_id})，状态码: {response.status_code}，响应: {response.text}")
		return False
	except Exception as e:
		print(f"Telegram图片发送异常({chat_id}): {e}")
		return False


def send_retention_to_tg(plat: str, pdf_path: str, excel_path: str, image_paths: list[str]):
	captions = [f"{plat} 首存留存 {date_str}", f"{plat} 首投留存 {date_str}"]

	for chat_id in SEND_ID:
		sent_any = False
		if SEND_AS_IMAGE and image_paths:
			for idx, image_path in enumerate(image_paths):
				caption = captions[idx] if idx < len(captions) else f"{plat} 留存报表 {date_str}"
				if SEND_IMAGE_AS_DOCUMENT:
					ok = send_telegram_document(chat_id, image_path, caption)
				else:
					ok = send_telegram_photo(chat_id, image_path, caption)
				sent_any = sent_any or ok

		if not sent_any and pdf_path and os.path.exists(pdf_path):
			sent_any = send_telegram_document(chat_id, pdf_path, f"{plat} 留存报表 {date_str}")

		if not sent_any:
			sent_any = send_telegram_document(chat_id, excel_path, f"{plat} 留存报表 {date_str}")

		if sent_any:
			print(f"{plat} 留存报表已发送到群 {chat_id}")
		else:
			print(f"{plat} 留存报表发送失败，群 {chat_id}")


def generate_retention_report(plat: str):
	print(f"[{plat}] 正在生成留存报表...")
	recharge_df = query_recharge_retention(plat)
	bet_df = query_bet_retention(plat)

	if recharge_df.empty and bet_df.empty:
		print(f"[{plat}] 没有留存数据，跳过")
		return

	recharge_df = add_summary_rows(recharge_df)
	bet_df = add_summary_rows(bet_df)

	excel_path = os.path.join(output_dir, f"{plat}_留存报表_{date_str}.xlsx")
	with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
		recharge_df.to_excel(writer, sheet_name="首充留存", index=False)
		bet_df.to_excel(writer, sheet_name="首投留存", index=False)

	style_retention_report(excel_path)
	print(f"[{plat}] Excel已保存: {excel_path}")

	pdf_path = export_excel_to_pdf(excel_path)
	image_paths = convert_pdf_pages_to_images(pdf_path, dpi=IMAGE_DPI, max_pages=2) if pdf_path else []

	if pdf_path:
		print(f"[{plat}] PDF已导出: {pdf_path}")
	if image_paths:
		print(f"[{plat}] 高清图已导出: {', '.join(image_paths)}")

	send_retention_to_tg(plat, pdf_path, excel_path, image_paths)


if __name__ == "__main__":
	# 先确认数据库可达
	with engine.connect():
		print("数据库连接成功")

	for platform in ["A8", "M9", "T1"]:
		try:
			generate_retention_report(platform)
		except Exception as e:
			print(f"[{platform}] 留存报表处理失败: {e}")

	print("留存报表生成完成！")
