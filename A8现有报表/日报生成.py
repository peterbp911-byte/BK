import os
import pandas as pd
from datetime import datetime, timedelta, date
from sqlalchemy import create_engine
import requests
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

TOKEN = '8718125099:AAH0lKs45qIyZUuAmWL8iE2NMh7H-lZ5hhI'
SEND_ID = ['-5255411609']
#SEND_ID = ['-5594167632'] # 测试用的
SEND_AS_IMAGE = True
SEND_IMAGE_AS_DOCUMENT = True
IMAGE_DPI = 360

today = datetime.today()
month_start_str = today.replace(day=1).strftime('%Y-%m-%d')
yesterday_str = (today - timedelta(days=1)).strftime('%Y-%m-%d')
date_str = today.strftime('%Y%m%d')

PERCENT_COLUMNS = {"首充率", "付费率", "提款占比", "出入比", "彩金占比", "游戏盈利率", "杀率"}
ONE_DECIMAL_COLUMNS = {"首充arppu", "arppu", "投充比"}

# 阈值预警（杀率使用字段“游戏盈利率”）
THRESHOLD_TABLE = {
    "A8": {"首充率": 0.6, "首充arppu": 400, "投充比": 7, "游戏盈利率": -0.1, "提款占比": 1},
    "M9": {"首充率": 0.1, "首充arppu": 400, "游戏盈利率": -0.1, "提款占比": 1},
    "T1": {"首充率": 0.6, "首充arppu": 400, "游戏盈利率": -0.1, "提款占比": 1},
}

THRESHOLD_RULES = {
    "首充率": "min",
    "首充arppu": "min",
    "投充比": "min",
    "游戏盈利率": "min",
    "杀率": "min",
    "提款占比": "max",
}


# 输出路径
output_dir = "C:/日报存档"
os.makedirs(output_dir, exist_ok=True)

engine = create_engine('postgresql+psycopg2://postgres:147258@localhost:5432')
if engine.connect():
    print("数据库连接成功")


def daily_report_data(plat):
    # 渠道报表SQL
    sql = f"""
    WITH game_day AS (
        SELECT
            "reportDate"::date AS d,
            COALESCE("betUserNum", 0) AS bet_user_num,
            COALESCE("betCoin", 0) AS bet_coin,
            COALESCE("winLoseCoin", 0) AS win_lose_coin
        FROM "public"."{plat}_game_statistics_report"
       
    ),
    agent_bonus_day AS (
        SELECT
            DATE(to_timestamp("operateTime" / 1000.0)) AS d,
            SUM(COALESCE(amount, 0)) AS agent_bonus_amt
        FROM "public"."{plat}_deposit_member_record"
        GROUP BY DATE(to_timestamp("operateTime" / 1000.0))
    ),
    revenue_day AS (
        SELECT
            "reportDate"::date AS d,
            COALESCE("boundAmt", 0) AS activity_bonus_amt,
            COALESCE("gameRebateAmt", 0) AS game_rebate_amt
        FROM "public"."{plat}_platfrom_daily_revenue"
    ),
    re_ret AS (
        SELECT
            "statisDate"::date AS d,
            COALESCE("retainRatio2", 0) / 100.0 AS retain_ratio_2,
            COALESCE("retainRatio3", 0) / 100.0 AS retain_ratio_3,
            COALESCE("retainRatio5", 0) / 100.0 AS retain_ratio_5,
            COALESCE("retainRatio7", 0) / 100.0 AS retain_ratio_7,
            COALESCE("retainRatio15", 0) / 100.0 AS retain_ratio_15,
            COALESCE("retainRatio30", 0) / 100.0 AS retain_ratio_30
        FROM "public"."{plat}_ReRetaintion"
    )
    SELECT
        od."statisDate"::date AS 日期,
        COALESCE(od."newUserNum", 0) AS 新增注册,
        COALESCE(gd.bet_user_num, 0) AS 日活,
        COALESCE(od."firstRechargeUserNum", 0) AS 首充人数,
        COALESCE(od."firstRechargeAmt", 0) AS 首充金额,
        ROUND(COALESCE(od."firstRechargeAmt", 0)::numeric / NULLIF(COALESCE(od."firstRechargeUserNum", 0), 0), 4) AS 首充arppu,
        ROUND(COALESCE(od."firstRechargeUserNum", 0)::numeric / NULLIF(COALESCE(od."newUserNum", 0), 0), 4) AS 首充率,
        COALESCE(od."rechargeUserNum", 0) AS 总充值人数,
        COALESCE(od."rechargeAmt", 0) AS 总充值金额,
        COALESCE(od."repeatRechargeUserNum", 0) AS 复充人数,
        COALESCE(od."repeatRechargeAmt", 0) AS 复充金额,
        ROUND(COALESCE(od."rechargeAmt", 0)::numeric / NULLIF(COALESCE(od."rechargeUserNum", 0), 0), 4) AS arppu,
        COALESCE(gd.bet_coin, 0) AS 投注金额,
        ROUND(COALESCE(gd.bet_coin, 0)::numeric / NULLIF(COALESCE(od."rechargeAmt", 0), 0), 4) AS 投充比,
        ROUND(COALESCE(od."rechargeUserNum", 0)::numeric / NULLIF(COALESCE(gd.bet_user_num, 0), 0), 4) AS 付费率,
        COALESCE(od."withdrawUserNum", 0) AS 提现人数,
        COALESCE(od."withdrawAmt", 0) AS 提现金额,
        ROUND(COALESCE(od."withdrawAmt", 0)::numeric / NULLIF(COALESCE(od."rechargeAmt", 0), 0), 4) AS 提款占比,
        ROUND(COALESCE(od."withdrawUserNum", 0)::numeric / NULLIF(COALESCE(od."rechargeUserNum", 0), 0), 4) AS 出入比,
        COALESCE(rv.activity_bonus_amt, 0) AS 活动彩金,
        COALESCE(rv.game_rebate_amt, 0) AS 游戏返水,
        COALESCE(abd.agent_bonus_amt, 0) AS 代理后台添加彩金,
        (COALESCE(rv.activity_bonus_amt, 0) + COALESCE(rv.game_rebate_amt, 0) + COALESCE(abd.agent_bonus_amt, 0)) AS 彩金总计,
        ROUND((COALESCE(rv.activity_bonus_amt, 0) + COALESCE(rv.game_rebate_amt, 0) + COALESCE(abd.agent_bonus_amt, 0))::numeric / NULLIF(COALESCE(gd.win_lose_coin, 0), 0), 4) AS 彩金占比,
        COALESCE(gd.win_lose_coin, 0) AS 公司输赢,
        ROUND(COALESCE(gd.win_lose_coin, 0)::numeric / NULLIF(COALESCE(gd.bet_coin, 0), 0), 4) AS 杀率,
        COALESCE(rr.retain_ratio_2, 0) AS "2日充值留存率",
        COALESCE(rr.retain_ratio_3, 0) AS "3日充值留存率",
        COALESCE(rr.retain_ratio_5, 0) AS "5日充值留存率",
        COALESCE(rr.retain_ratio_7, 0) AS "7日充值留存率",
        COALESCE(rr.retain_ratio_15, 0) AS "14日充值留存率"
    FROM "public"."{plat}_opration_data" od
    LEFT JOIN game_day gd ON gd.d = od."statisDate"::date
    LEFT JOIN revenue_day rv ON rv.d = od."statisDate"::date
    LEFT JOIN agent_bonus_day abd ON abd.d = od."statisDate"::date
    LEFT JOIN re_ret rr ON rr.d = od."statisDate"::date
    WHERE od."statisDate" BETWEEN '{month_start_str}'::date AND '{yesterday_str}'::date
    ORDER BY od."statisDate" DESC

    """


    # 创建文件名
    print(f"[A8-每日平台数据]正在处理数据中...")
    df = pd.read_sql(sql, engine)

    if not df.empty:
        def _to_num_series(series: pd.Series) -> pd.Series:
            # Support numeric strings and percentage-like strings.
            cleaned = series.astype(str).str.replace('%', '', regex=False).str.replace(',', '', regex=False)
            return pd.to_numeric(cleaned, errors='coerce')

        profit_col = "公司输赢" if "公司输赢" in df.columns else "三方游戏盈亏"
        kill_rate_col = "杀率" if "杀率" in df.columns else "游戏盈利率"

        ratio_map = {
            "首充arppu": ("首充金额", "首充人数"),
            "首充率": ("首充人数", "新增注册"),
            "arppu": ("总充值金额", "总充值人数"),
            "投充比": ("投注金额", "总充值金额"),
            "付费率": ("总充值人数", "日活"),
            "提款占比": ("提现金额", "总充值金额"),
            "出入比": ("提现人数", "总充值人数"),
            "彩金占比": ("彩金总计", profit_col),
            kill_rate_col: (profit_col, "投注金额"),
        }

        numeric_cols = [c for c in df.columns if c != "日期"]
        numeric_data = {c: _to_num_series(df[c]) for c in numeric_cols}
        day_count = len(df)

        summary_row = {c: "" for c in df.columns}
        average_row = {c: "" for c in df.columns}
        summary_row["日期"] = "汇总"
        average_row["日期"] = "平均"

        for c in numeric_cols:
            if c in ratio_map:
                continue
            s = numeric_data[c].sum(skipna=True)
            summary_row[c] = round(float(s), 4)
            average_row[c] = round(float(s) / day_count, 4) if day_count else 0

        def _safe_div(num: float, den: float) -> float:
            if den == 0:
                return 0
            return round(num / den, 4)

        for ratio_col, (num_col, den_col) in ratio_map.items():
            if ratio_col in df.columns and num_col in numeric_data and den_col in numeric_data:
                num_sum = float(numeric_data[num_col].sum(skipna=True))
                den_sum = float(numeric_data[den_col].sum(skipna=True))
                ratio_val = _safe_div(num_sum, den_sum)
                summary_row[ratio_col] = ratio_val
                average_row[ratio_col] = ratio_val
            elif ratio_col in numeric_data:
                # Fallback when source numerator/denominator columns are renamed or absent.
                ratio_mean = numeric_data[ratio_col].mean(skipna=True)
                if pd.notna(ratio_mean):
                    summary_row[ratio_col] = round(float(ratio_mean), 4)
                    average_row[ratio_col] = round(float(ratio_mean), 4)

        # Retention ratios have no numerator/denominator columns in this table, use arithmetic mean.
        for c in df.columns:
            if c.endswith("留存率") and c not in ratio_map:
                c_mean = numeric_data[c].mean(skipna=True) if c in numeric_data else None
                if pd.notna(c_mean):
                    summary_row[c] = round(float(c_mean), 4)
                    average_row[c] = round(float(c_mean), 4)

        df = pd.concat([pd.DataFrame([summary_row, average_row]), df], ignore_index=True)

    return df


def _to_float(value) -> float:
    if value is None:
        return 0.0
    if isinstance(value, str):
        value = value.replace('%', '').replace(',', '').strip()
    num = pd.to_numeric(value, errors='coerce')
    return float(num) if pd.notna(num) else 0.0


def _format_people(value) -> str:
    return f"{int(round(_to_float(value)))}"


def _format_amount(value) -> str:
    num = _to_float(value)
    if abs(num - round(num)) < 1e-9:
        return f"{int(round(num))}"
    return f"{num:.2f}"


def _format_percent(value) -> str:
    num = _to_float(value)
    return f"{num * 100:.2f}%"


def _pick_detail_row(df: pd.DataFrame):
    if df.empty or "日期" not in df.columns:
        return None

    y_date = datetime.strptime(yesterday_str, "%Y-%m-%d").date()

    def _as_date(value):
        if isinstance(value, date):
            return value
        if hasattr(value, "date"):
            try:
                return value.date()
            except Exception:
                return None
        if isinstance(value, str):
            try:
                return pd.to_datetime(value).date()
            except Exception:
                return None
        return None

    for _, row in df.iterrows():
        date_val = _as_date(row.get("日期"))
        if date_val == y_date:
            return row

    for _, row in df.iterrows():
        date_val = _as_date(row.get("日期"))
        if date_val is not None:
            return row

    return None


def build_daily_summary_text(df: pd.DataFrame, plat: str) -> str:
    row = _pick_detail_row(df)
    summary_date = today - timedelta(days=1)
    display_date = f"{summary_date.year}.{summary_date.month}.{summary_date.day}"
    if row is None:
        return f"{display_date}  {plat}数据\n暂无可用数据"

    register_num = _format_people(row.get("新增注册"))
    first_recharge_users = _format_people(row.get("首充人数"))
    first_recharge_amt = _format_amount(row.get("首充金额"))
    recharge_users = _format_people(row.get("总充值人数"))
    recharge_amt = _format_amount(row.get("总充值金额"))
    withdraw_amt = _format_amount(row.get("提现金额"))

    charge_minus_withdraw = _to_float(row.get("总充值金额")) - _to_float(row.get("提现金额"))
    charge_minus_withdraw_text = _format_amount(charge_minus_withdraw)

    bet_users = _format_people(row.get("日活"))
    profit_col = "公司输赢" if "公司输赢" in df.columns else "三方游戏盈亏"
    game_profit = _format_amount(row.get(profit_col))
    kill_rate_col = "杀率" if "杀率" in df.columns else "游戏盈利率"
    kill_rate = _format_percent(row.get(kill_rate_col))

    summary_lines = [
        f"{display_date}  {plat}数据",
        f"新增注册：{register_num}人",
        f"首充人数：{first_recharge_users}人",
        f"首充金额：{first_recharge_amt}",
        f"充值人数：{recharge_users}人",
        f"充值：{recharge_amt}",
        f"提现：{withdraw_amt}",
        f"充提差：{charge_minus_withdraw_text}",
        "—————————",
        f"投注人数：{bet_users}",
        f"三方游戏盈亏：{game_profit}",
        f"游戏盈利率：{kill_rate}",
    ]
    return "\n".join(summary_lines)


def save_daily_report(df: pd.DataFrame, plat: str):
    output_path = os.path.join(output_dir, f"{plat}_日报_{date_str}.xlsx")
    df.to_excel(output_path, index=False)
    style_daily_report(output_path, plat)
    summary_text = build_daily_summary_text(df, plat)
    print(summary_text)
    pdf_path = export_excel_to_pdf(output_path)
    image_path = None
    print(f"{plat} 报表已保存: {output_path}")
    if pdf_path:
        print(f"{plat} PDF已导出: {pdf_path}")
        image_path = convert_pdf_to_image(pdf_path, dpi=IMAGE_DPI)
        if image_path:
            print(f"{plat} 高清图片已导出: {image_path}")
    else:
        print(f"{plat} PDF导出失败，将回退发送Excel")

    # 发送到 Telegram（优先发送高清图片，其次 PDF，最后 Excel）
    send_path = output_path
    caption = f"{plat} 日报 {date_str}"
    for chat_id in SEND_ID:
        ok = False
        sent_path = None
        if SEND_AS_IMAGE and image_path and os.path.exists(image_path):
            if SEND_IMAGE_AS_DOCUMENT:
                ok = send_telegram_document(chat_id=chat_id, file_path=image_path, caption=caption)
            else:
                ok = send_telegram_photo(chat_id=chat_id, image_path=image_path, caption=caption)
            if ok:
                sent_path = image_path
                mode_text = "图片文件" if SEND_IMAGE_AS_DOCUMENT else "图片"
                print(f"{plat} {mode_text}已发送到群 {chat_id}: {os.path.basename(image_path)}")

        if not ok:
            if pdf_path and os.path.exists(pdf_path):
                send_path = pdf_path
            ok = send_telegram_document(chat_id=chat_id, file_path=send_path, caption=caption)
            if ok:
                sent_path = send_path

        if ok:
            print(f"{plat} 报表已发送到群 {chat_id}: {os.path.basename(sent_path) if sent_path else '-'}")
        else:
            print(f"{plat} 报表发送失败，群 {chat_id}: {os.path.basename(send_path)}")

        txt_ok = send_telegram_message(chat_id=chat_id, text=summary_text)
        if txt_ok:
            print(f"{plat} 摘要已发送到群 {chat_id}")
        else:
            print(f"{plat} 摘要发送失败，群 {chat_id}")


def send_telegram_message(chat_id: str, text: str) -> bool:
    send_url = f"https://api.telegram.org/bot{TOKEN}/sendMessage"
    try:
        response = requests.post(
            send_url,
            data={"chat_id": chat_id, "text": text},
            timeout=60,
        )
        if response.status_code == 200:
            return True
        print(f"Telegram消息发送失败({chat_id})，状态码: {response.status_code}，响应: {response.text}")
        return False
    except Exception as e:
        print(f"Telegram消息发送异常({chat_id}): {e}")
        return False


def send_telegram_document(chat_id: str, file_path: str, caption: str = "") -> bool:
    if not os.path.exists(file_path):
        print(f"待发送文件不存在: {file_path}")
        return False

    send_url = f"https://api.telegram.org/bot{TOKEN}/sendDocument"
    try:
        with open(file_path, "rb") as f:
            response = requests.post(
                send_url,
                data={"chat_id": chat_id, "caption": caption},
                files={"document": f},
                timeout=60,
            )
        if response.status_code == 200:
            return True
        print(f"Telegram发送失败({chat_id})，状态码: {response.status_code}，响应: {response.text}")
        return False
    except Exception as e:
        print(f"Telegram发送异常({chat_id}): {e}")
        return False


def send_telegram_photo(chat_id: str, image_path: str, caption: str = "") -> bool:
    if not os.path.exists(image_path):
        print(f"待发送图片不存在: {image_path}")
        return False

    send_url = f"https://api.telegram.org/bot{TOKEN}/sendPhoto"
    try:
        with open(image_path, "rb") as f:
            response = requests.post(
                send_url,
                data={"chat_id": chat_id, "caption": caption},
                files={"photo": f},
                timeout=60,
            )
        if response.status_code == 200:
            return True
        print(f"Telegram图片发送失败({chat_id})，状态码: {response.status_code}，响应: {response.text}")
        return False
    except Exception as e:
        print(f"Telegram图片发送异常({chat_id}): {e}")
        return False


def convert_pdf_to_image(pdf_path: str, dpi: int = 240):
    if not os.path.exists(pdf_path):
        return None

    try:
        import fitz

        zoom = dpi / 72
        image_path = os.path.splitext(pdf_path)[0] + ".png"
        doc = fitz.open(pdf_path)
        try:
            if doc.page_count < 1:
                return None
            page = doc.load_page(0)
            matrix = fitz.Matrix(zoom, zoom)

            # 按内容区域裁剪，避免导出整页造成下方大片留白。
            blocks = page.get_text("blocks")
            if blocks:
                x0 = min(b[0] for b in blocks)
                y0 = min(b[1] for b in blocks)
                x1 = max(b[2] for b in blocks)
                y1 = max(b[3] for b in blocks)
                margin = 12  # point
                clip = fitz.Rect(
                    max(page.rect.x0, x0 - margin),
                    max(page.rect.y0, y0 - margin),
                    min(page.rect.x1, x1 + margin),
                    min(page.rect.y1, y1 + margin),
                )
            else:
                clip = page.rect

            pix = page.get_pixmap(matrix=matrix, clip=clip, alpha=False)
            pix.save(image_path)
        finally:
            doc.close()
        return image_path
    except Exception as e:
        print(f"PDF转图片失败: {e}")
        return None


def _is_number(value) -> bool:
    return isinstance(value, (int, float)) and pd.notna(value)


def _check_threshold(value: float, threshold: float, rule: str) -> bool:
    if rule == "min":
        return value < threshold
    if rule == "max":
        return value > threshold
    return False


def style_daily_report(output_path: str, plat: str):
    wb = load_workbook(output_path)
    ws = wb.active

    # 插入标题行
    ws.insert_rows(1)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ws.max_column)
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = f"{plat} 运营日报（{yesterday_str}）"

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    title_fill = PatternFill("solid", fgColor="163E63")
    title_font = Font(color="FFFFFF", bold=True, size=14)
    summary_fill = PatternFill("solid", fgColor="E7E6E6")
    warn_fill = PatternFill("solid", fgColor="F8CBAD")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    title_cell.fill = title_fill
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.border = border

    # 表头样式
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=2, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    col_name_to_idx = {ws.cell(row=2, column=i).value: i for i in range(1, ws.max_column + 1)}
    yesterday_date = datetime.strptime(yesterday_str, "%Y-%m-%d").date()

    # 数据样式与数字格式
    date_rows = []
    for row_idx in range(3, ws.max_row + 1):
        row_label = ws.cell(row=row_idx, column=1).value
        is_summary = str(row_label) in {"汇总", "平均"}

        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            col_name = ws.cell(row=2, column=col_idx).value
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")

            if is_summary:
                cell.fill = summary_fill
                cell.font = Font(bold=True)

            if _is_number(cell.value) and col_name:
                if col_name in PERCENT_COLUMNS or str(col_name).endswith("留存率"):
                    cell.number_format = "0.0%"
                elif col_name in ONE_DECIMAL_COLUMNS:
                    cell.number_format = "0.0"
                else:
                    cell.number_format = "#,##0"

        if hasattr(row_label, "date"):
            date_rows.append((row_idx, row_label.date()))

    # 阈值预警：优先检查昨天数据，如未找到则检查最新日期行
    threshold_values = THRESHOLD_TABLE.get(plat, {})
    target_row = None
    for r, d in date_rows:
        if d == yesterday_date:
            target_row = r
            break
    if target_row is None and date_rows:
        target_row = date_rows[0][0]

    warning_records = []
    if target_row and threshold_values:
        check_date = ws.cell(row=target_row, column=1).value
        for metric, threshold in threshold_values.items():
            metric_alias = {"游戏盈利率": "杀率", "三方游戏盈亏": "公司输赢"}
            col_idx = col_name_to_idx.get(metric) or col_name_to_idx.get(metric_alias.get(metric, ""))
            if not col_idx:
                continue
            metric_cell = ws.cell(row=target_row, column=col_idx)
            metric_value = metric_cell.value
            if not _is_number(metric_value):
                continue

            rule = THRESHOLD_RULES.get(metric)
            if rule and _check_threshold(float(metric_value), float(threshold), rule):
                metric_cell.fill = warn_fill
                warning_records.append({
                    "平台": plat,
                    "日期": check_date,
                    "指标": metric,
                    "当前值": float(metric_value),
                    "规则": "低于触发" if rule == "min" else "高于触发",
                })

    # 预警明细页
    if "阈值预警" in wb.sheetnames:
        del wb["阈值预警"]
    warn_ws = wb.create_sheet("阈值预警")
    warn_headers = ["平台", "日期", "指标", "当前值", "规则"]
    warn_ws.append(warn_headers)
    for c in range(1, len(warn_headers) + 1):
        h_cell = warn_ws.cell(row=1, column=c)
        h_cell.fill = header_fill
        h_cell.font = header_font
        h_cell.alignment = Alignment(horizontal="center", vertical="center")
        h_cell.border = border

    if warning_records:
        for rec in warning_records:
            warn_ws.append([rec["平台"], rec["日期"], rec["指标"], rec["当前值"], rec["规则"]])
    else:
        warn_ws.append([plat, yesterday_str, "-", "无触发预警", "-"])

    for row_idx in range(2, warn_ws.max_row + 1):
        for col_idx in range(1, warn_ws.max_column + 1):
            cell = warn_ws.cell(row=row_idx, column=col_idx)
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # 自适应列宽、行高
    for active_ws in [ws, warn_ws]:
        for col_idx in range(1, active_ws.max_column + 1):
            max_len = 0
            for row_idx in range(1, active_ws.max_row + 1):
                value = active_ws.cell(row=row_idx, column=col_idx).value
                if value is None:
                    continue
                text = str(value)
                max_len = max(max_len, len(text))
            active_ws.column_dimensions[get_column_letter(col_idx)].width = min(max(12, max_len + 4), 36)

        if active_ws == ws:
            active_ws.row_dimensions[1].height = 30
            active_ws.row_dimensions[2].height = 34
            for row_idx in range(3, active_ws.max_row + 1):
                active_ws.row_dimensions[row_idx].height = 20
        else:
            active_ws.row_dimensions[1].height = 34
            for row_idx in range(2, active_ws.max_row + 1):
                active_ws.row_dimensions[row_idx].height = 20

    # 横版与打印缩放
    ws.freeze_panes = "A3"
    ws.sheet_view.zoomScale = 90
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_area = ws.dimensions

    warn_ws.freeze_panes = "A2"
    warn_ws.page_setup.orientation = warn_ws.ORIENTATION_LANDSCAPE
    warn_ws.page_setup.paperSize = warn_ws.PAPERSIZE_A4
    warn_ws.sheet_properties.pageSetUpPr.fitToPage = True
    warn_ws.page_setup.fitToWidth = 1
    warn_ws.page_setup.fitToHeight = 0
    warn_ws.print_area = warn_ws.dimensions

    wb.save(output_path)


def export_excel_to_pdf(excel_path: str):
    pdf_path = os.path.splitext(excel_path)[0] + ".pdf"
    excel_app = None
    workbook = None
    try:
        import win32com.client

        excel_app = win32com.client.DispatchEx("Excel.Application")
        excel_app.Visible = False
        excel_app.DisplayAlerts = False
        workbook = excel_app.Workbooks.Open(os.path.abspath(excel_path))

        # 二次保险：强制每个工作表按 1 页宽缩放，确保所有列都在一页宽度内。
        for sheet in workbook.Worksheets:
            sheet.PageSetup.Zoom = False
            sheet.PageSetup.Orientation = 2  # xlLandscape
            sheet.PageSetup.FitToPagesWide = 1
            sheet.PageSetup.FitToPagesTall = False

        # 仅导出第一页主报表，不打印第二页阈值预警页。
        workbook.Worksheets(1).ExportAsFixedFormat(0, os.path.abspath(pdf_path))
        return pdf_path
    except Exception as e:
        print(f"PDF导出失败（可先安装 pywin32 或检查本机Excel）: {e}")
        return None
    finally:
        if workbook is not None:
            workbook.Close(False)
        if excel_app is not None:
            excel_app.Quit()


if __name__ == "__main__":
    platform_dfs = {}
    for plat in ['A8', 'M9', 'T1']:
        df = daily_report_data(plat)
        platform_dfs[plat] = df
        save_daily_report(df, plat)

    print("日报表生成完成！")
